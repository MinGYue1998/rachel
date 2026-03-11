from fastapi import APIRouter, Query

from app.controllers.education import report_controller
from app.core.dependency import DependAuth
from app.schemas.base import Success

router = APIRouter(dependencies=[DependAuth])


@router.get("/monthly", summary="月度报表")
async def get_monthly_report(
    year: int = Query(None, description="年份"),
    month: int = Query(None, description="月份"),
):
    report = await report_controller.get_monthly_report(year, month)
    return Success(data=report)


@router.get("/student-detail", summary="学生明细报表")
async def get_student_detail_report(
    student_id: int = Query(None, description="学生ID"),
    course_id: int = Query(None, description="课程ID"),
    year: int = Query(None, description="年份"),
    month: int = Query(None, description="月份"),
):
    report = await report_controller.get_student_detail_report(
        student_id=student_id,
        course_id=course_id,
        year=year,
        month=month,
    )
    return Success(data=report)


@router.get("/export", summary="导出报表")
async def export_report(
    report_type: str = Query(..., description="报表类型: monthly/student-detail"),
    year: int = Query(None, description="年份"),
    month: int = Query(None, description="月份"),
    student_id: int = Query(None, description="学生ID"),
):
    from fastapi.responses import StreamingResponse
    from io import BytesIO, StringIO
    
    # 获取数据
    if report_type == "monthly":
        data = await report_controller.get_monthly_report(year, month)
        headers = ["月份", "课时费合计", "缴费合计", "上课次数", "学生人数"]
        rows = [[d["month"], d["total_class_fee"], d["total_payment"], d["class_count"], d["student_count"]] for d in data]
    else:
        data = await report_controller.get_student_detail_report(student_id=student_id, year=year, month=month)
        headers = ["学生ID", "学生姓名", "课程", "上课次数", "总课时", "总费用", "已缴费", "欠费"]
        rows = [[d["student_id"], d["student_name"], d["course_name"] or "", d["class_count"], d["total_hours"], d["total_fee"], d["total_paid"], d["balance"]] for d in data]
    
    # 生成 CSV
    import csv
    output = BytesIO()
    # 添加 BOM 以支持中文
    output.write(b'\xef\xbb\xbf')
    
    # 使用 StringIO 写入 CSV
    text_output = StringIO()
    writer = csv.writer(text_output)
    writer.writerow(headers)
    writer.writerows(rows)
    
    output.write(text_output.getvalue().encode('utf-8'))
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={report_type}_report.csv"
        }
    )


@router.get("/export-excel", summary="导出报表Excel")
async def export_report_excel(
    report_type: str = Query(..., description="报表类型: monthly/student-detail"),
    year: int = Query(None, description="年份"),
    month: int = Query(None, description="月份"),
    student_id: int = Query(None, description="学生ID"),
    course_id: int = Query(None, description="课程ID"),
):
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # 获取数据
    if report_type == "monthly":
        data = await report_controller.get_monthly_report(year, month)
        headers = ["月份", "课时费合计", "缴费合计", "上课次数", "学生人数"]
        rows = [[d["month"], d["total_class_fee"], d["total_payment"], d["class_count"], d["student_count"]] for d in data]
        sheet_name = "月度汇总报表"
    else:
        data = await report_controller.get_student_detail_report(
            student_id=student_id, course_id=course_id, year=year, month=month
        )
        headers = ["学生ID", "学生姓名", "课程", "上课次数", "总课时", "总费用", "优惠金额", "已缴费", "欠费"]
        rows = [[
            d["student_id"], d["student_name"], d["course_name"] or "", 
            d["class_count"], d["total_hours"], d["total_fee"], 
            d["total_discount"], d["total_paid"], d["balance"]
        ] for d in data]
        sheet_name = "学生明细报表"
    
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 设置样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # 写入数据
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.5
        ws.column_dimensions[column].width = max(adjusted_width, 10)
    
    # 保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={report_type}_report.xlsx"
        }
    )
